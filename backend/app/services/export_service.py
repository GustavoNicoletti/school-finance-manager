from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FILL = PatternFill("solid", fgColor="0F172A")
SUBTITLE_FILL = PatternFill("solid", fgColor="E2E8F0")
FILTER_LABEL_FILL = PatternFill("solid", fgColor="DBEAFE")
STATUS_FILLS = {
    "pendente": PatternFill("solid", fgColor="FEF3C7"),
    "atrasado": PatternFill("solid", fgColor="FEE2E2"),
    "pago": PatternFill("solid", fgColor="DCFCE7"),
    "cancelado": PatternFill("solid", fgColor="E5E7EB"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


@dataclass(slots=True)
class ExcelReportSection:
    label: str
    value: object


@dataclass(slots=True)
class ExcelReportConfig:
    filename: str
    sheet_name: str
    title: str
    headers: list[str]
    rows: list[list[object]]
    filters: list[ExcelReportSection] = field(default_factory=list)
    summary: list[ExcelReportSection] = field(default_factory=list)
    currency_columns: set[int] = field(default_factory=set)
    date_columns: set[int] = field(default_factory=set)
    integer_columns: set[int] = field(default_factory=set)
    status_column: int | None = None


@dataclass(slots=True)
class PdfReportConfig:
    filename: str
    title: str
    headers: list[str]
    rows: list[list[object]]
    filters: list[ExcelReportSection] = field(default_factory=list)
    summary: list[ExcelReportSection] = field(default_factory=list)


def _normalize_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    return value


def _display_value(value: object) -> str:
    if value in (None, ""):
        return "-"
    if isinstance(value, Decimal):
        return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if isinstance(value, float):
        return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)


def _paragraph(value: object, style: ParagraphStyle, *, bold: bool = False) -> Paragraph:
    text = escape(_display_value(value))
    if bold:
        text = f"<b>{text}</b>"
    return Paragraph(text, style)


def _set_cell_value(cell, value: object) -> None:
    normalized = _normalize_value(value)
    cell.value = normalized if normalized not in ("", None) else None


def _section_rows(start_row: int, sections: list[ExcelReportSection], title: str, worksheet) -> int:
    if not sections:
        return start_row

    worksheet.cell(row=start_row, column=1, value=title)
    worksheet.cell(row=start_row, column=1).font = Font(bold=True, color="0F172A")
    worksheet.cell(row=start_row, column=1).fill = SUBTITLE_FILL
    worksheet.cell(row=start_row, column=1).border = THIN_BORDER
    worksheet.cell(row=start_row, column=2).border = THIN_BORDER

    current_row = start_row + 1
    for item in sections:
        label_cell = worksheet.cell(row=current_row, column=1, value=item.label)
        value_cell = worksheet.cell(row=current_row, column=2)
        _set_cell_value(value_cell, item.value)
        label_cell.font = Font(bold=True, color="1E293B")
        label_cell.fill = FILTER_LABEL_FILL
        label_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER
        current_row += 1
    return current_row


def _apply_data_formats(config: ExcelReportConfig, worksheet, header_row: int, first_data_row: int, last_data_row: int) -> None:
    if last_data_row < first_data_row:
        return

    for row in worksheet.iter_rows(min_row=first_data_row, max_row=last_data_row):
        for cell in row:
            column_index = cell.column
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")

            if column_index in config.currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
            elif column_index in config.date_columns and cell.value:
                cell.number_format = "DD/MM/YYYY"
            elif column_index in config.integer_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0"

            if config.status_column and column_index == config.status_column and isinstance(cell.value, str):
                status_fill = STATUS_FILLS.get(cell.value.lower())
                if status_fill:
                    cell.fill = status_fill
                    cell.font = Font(bold=True, color="334155")

    for cell in worksheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 36)


def build_excel_report(config: ExcelReportConfig) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = config.sheet_name
    worksheet.sheet_view.showGridLines = False

    column_count = max(len(config.headers), 2)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = worksheet.cell(row=1, column=1, value=config.title)
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = THIN_BORDER

    generated_at = worksheet.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    generated_at.font = Font(italic=True, color="475569")

    current_row = 4
    current_row = _section_rows(current_row, config.filters, "Filtros aplicados", worksheet)
    if config.summary:
        current_row += 1 if config.filters else 0
        current_row = _section_rows(current_row, config.summary, "Resumo", worksheet)

    current_row += 1
    header_row = current_row
    for index, header in enumerate(config.headers, start=1):
        worksheet.cell(row=header_row, column=index, value=header)

    first_data_row = header_row + 1
    for row_index, row in enumerate(config.rows, start=first_data_row):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            _set_cell_value(cell, value)

    last_data_row = header_row + len(config.rows)
    _apply_data_formats(config, worksheet, header_row, first_data_row, last_data_row)

    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(config.headers))}{max(last_data_row, header_row)}"
    worksheet.freeze_panes = f"A{first_data_row}"
    worksheet.row_dimensions[1].height = 24
    _autofit_columns(worksheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf_report(config: PdfReportConfig) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=6,
    )
    section_style = ParagraphStyle(
        "ReportSection",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1E293B"),
        spaceBefore=8,
        spaceAfter=4,
    )
    body_style = ParagraphStyle(
        "ReportBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0F172A"),
    )

    story = [
        Paragraph(config.title, title_style),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", body_style),
        Spacer(1, 8),
    ]

    if config.filters:
        story.append(Paragraph("Filtros aplicados", section_style))
        filter_data = [[_paragraph("Campo", body_style, bold=True), _paragraph("Valor", body_style, bold=True)]]
        filter_data.extend([[_paragraph(item.label, body_style), _paragraph(item.value, body_style)] for item in config.filters])
        filter_table = Table(filter_data, colWidths=[55 * mm, 190 * mm], repeatRows=1)
        filter_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DBEAFE")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.extend([filter_table, Spacer(1, 8)])

    if config.summary:
        story.append(Paragraph("Resumo", section_style))
        summary_data = [[_paragraph("Indicador", body_style, bold=True), _paragraph("Valor", body_style, bold=True)]]
        summary_data.extend([[_paragraph(item.label, body_style), _paragraph(item.value, body_style)] for item in config.summary])
        summary_table = Table(summary_data, colWidths=[65 * mm, 70 * mm], repeatRows=1)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.extend([summary_table, Spacer(1, 10)])

    table_data = [
        [_paragraph(header, body_style, bold=True) for header in config.headers],
        *[[_paragraph(value, body_style) for value in row] for row in config.rows],
    ]
    page_width = landscape(A4)[0] - document.leftMargin - document.rightMargin
    column_width = page_width / max(len(config.headers), 1)
    report_table = Table(table_data, colWidths=[column_width] * len(config.headers), repeatRows=1)
    report_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(report_table)

    document.build(story)
    return output.getvalue()
