from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from app.models.finance import Payable, PaymentStatus, Receivable, ReceivableType
from app.models.student import Student, StudentStatus
from app.models.user import Role
from .conftest import auth_headers, create_user


def workbook_rows(response) -> list[tuple[object, ...]]:
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def test_delinquency_export_csv_returns_expected_content(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    student = Student(full_name="Aluno Inadimplente", status=StudentStatus.ACTIVE, class_name="2 Ano")
    receivable = Receivable(
        student=student,
        description="Mensalidade 09/2026",
        amount=Decimal("700.00"),
        paid_amount=Decimal("0.00"),
        due_date=date(2026, 9, 5),
        status=PaymentStatus.OVERDUE,
        type=ReceivableType.TUITION,
    )
    db.add_all([student, receivable])
    db.commit()

    response = client.get(
        "/api/finance/delinquency/export.csv",
        params={"as_of_date": "2026-09-10"},
        headers=auth_headers(client, "finance@example.com"),
    )

    body = response.content.decode("utf-8")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="inadimplencia.csv"'
    assert body.startswith("\ufeff")
    assert "aluno;turma;telefone;parcelas;mais_antiga;valor_em_aberto" in body
    assert "Aluno Inadimplente;2 Ano;;1;2026-09-05;700.00" in body


def test_delinquency_export_xlsx_returns_formatted_workbook(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    student = Student(full_name="Aluno Inadimplente", status=StudentStatus.ACTIVE, class_name="2 Ano")
    receivable = Receivable(
        student=student,
        description="Mensalidade 09/2026",
        amount=Decimal("700.00"),
        paid_amount=Decimal("0.00"),
        due_date=date(2026, 9, 5),
        status=PaymentStatus.OVERDUE,
        type=ReceivableType.TUITION,
    )
    db.add_all([student, receivable])
    db.commit()

    response = client.get(
        "/api/finance/delinquency/export.xlsx",
        params={"as_of_date": "2026-09-10"},
        headers=auth_headers(client, "finance@example.com"),
    )

    rows = workbook_rows(response)
    flattened = [value for row in rows for value in row if value is not None]
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="inadimplencia.xlsx"'
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]
    assert "Relatorio de Inadimplencia" in flattened
    assert "Aluno Inadimplente" in flattened
    assert 700 in flattened


def test_receivables_export_csv_respects_filters(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    student_a = Student(full_name="Alice Rocha", status=StudentStatus.ACTIVE)
    student_b = Student(full_name="Bruno Costa", status=StudentStatus.ACTIVE)
    db.add_all([student_a, student_b])
    db.flush()
    db.add_all(
        [
            Receivable(
                student_id=student_a.id,
                description="Mensalidade Alice",
                amount=Decimal("500.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 9, 5),
                status=PaymentStatus.PENDING,
                type=ReceivableType.TUITION,
            ),
            Receivable(
                student_id=student_b.id,
                description="Mensalidade Bruno",
                amount=Decimal("650.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 9, 5),
                status=PaymentStatus.PENDING,
                type=ReceivableType.EXTRA_FEE,
            ),
        ]
    )
    db.commit()

    response = client.get(
        "/api/finance/receivables/export.csv",
        params={"search": "Alice"},
        headers=auth_headers(client, "finance@example.com"),
    )

    body = response.content.decode("utf-8")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="contas_a_receber.csv"'
    assert "Alice Rocha;Mensalidade Alice" in body
    assert "Bruno Costa;Mensalidade Bruno" not in body


def test_receivables_export_xlsx_respects_filters(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    student_a = Student(full_name="Alice Rocha", status=StudentStatus.ACTIVE)
    student_b = Student(full_name="Bruno Costa", status=StudentStatus.ACTIVE)
    db.add_all([student_a, student_b])
    db.flush()
    db.add_all(
        [
            Receivable(
                student_id=student_a.id,
                description="Mensalidade Alice",
                amount=Decimal("500.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 9, 5),
                status=PaymentStatus.PENDING,
                type=ReceivableType.TUITION,
            ),
            Receivable(
                student_id=student_b.id,
                description="Mensalidade Bruno",
                amount=Decimal("650.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 9, 5),
                status=PaymentStatus.PENDING,
                type=ReceivableType.EXTRA_FEE,
            ),
        ]
    )
    db.commit()

    response = client.get(
        "/api/finance/receivables/export.xlsx",
        params={"search": "Alice"},
        headers=auth_headers(client, "finance@example.com"),
    )

    rows = workbook_rows(response)
    flattened = [value for row in rows for value in row if value is not None]
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="contas_a_receber.xlsx"'
    assert "Alice Rocha" in flattened
    assert "Mensalidade Alice" in flattened
    assert "Bruno Costa" not in flattened


def test_payables_export_csv_respects_filters(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    db.add_all(
        [
            Payable(
                description="Aluguel setembro",
                category="estrutura",
                amount=Decimal("1200.00"),
                due_date=date(2026, 9, 10),
                status=PaymentStatus.PENDING,
                supplier="Imobiliaria Centro",
            ),
            Payable(
                description="Internet setembro",
                category="utilidades",
                amount=Decimal("300.00"),
                due_date=date(2026, 9, 12),
                status=PaymentStatus.PENDING,
                supplier="Operadora Conecta",
            ),
        ]
    )
    db.commit()

    response = client.get(
        "/api/finance/payables/export.csv",
        params={"category": "estrutura"},
        headers=auth_headers(client, "finance@example.com"),
    )

    body = response.content.decode("utf-8")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="contas_a_pagar.csv"'
    assert "Aluguel setembro;estrutura;Imobiliaria Centro" in body
    assert "Internet setembro;utilidades;Operadora Conecta" not in body


def test_payables_export_xlsx_respects_filters(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    db.add_all(
        [
            Payable(
                description="Aluguel setembro",
                category="estrutura",
                amount=Decimal("1200.00"),
                due_date=date(2026, 9, 10),
                status=PaymentStatus.PENDING,
                supplier="Imobiliaria Centro",
            ),
            Payable(
                description="Internet setembro",
                category="utilidades",
                amount=Decimal("300.00"),
                due_date=date(2026, 9, 12),
                status=PaymentStatus.PENDING,
                supplier="Operadora Conecta",
            ),
        ]
    )
    db.commit()

    response = client.get(
        "/api/finance/payables/export.xlsx",
        params={"category": "estrutura"},
        headers=auth_headers(client, "finance@example.com"),
    )

    rows = workbook_rows(response)
    flattened = [value for row in rows for value in row if value is not None]
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="contas_a_pagar.xlsx"'
    assert "Aluguel setembro" in flattened
    assert "Imobiliaria Centro" in flattened
    assert "Internet setembro" not in flattened


def test_cash_flow_export_csv_returns_period_entries(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    student = Student(full_name="Aluno Fluxo Export", status=StudentStatus.ACTIVE)
    receivable = Receivable(
        student=student,
        description="Mensalidade exportacao",
        amount=Decimal("800.00"),
        paid_amount=Decimal("800.00"),
        due_date=date(2026, 9, 5),
        payment_date=date(2026, 9, 6),
        status=PaymentStatus.PAID,
        type=ReceivableType.TUITION,
    )
    payable = Payable(
        description="Fornecedor exportacao",
        category="servicos",
        amount=Decimal("250.00"),
        due_date=date(2026, 9, 6),
        payment_date=date(2026, 9, 6),
        status=PaymentStatus.PAID,
    )
    db.add_all([student, receivable, payable])
    db.commit()

    response = client.get(
        "/api/finance/cash-flow/export.csv",
        params={"start_date": "2026-09-05", "end_date": "2026-09-06"},
        headers=auth_headers(client, "finance@example.com"),
    )

    body = response.content.decode("utf-8")
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="fluxo_caixa.csv"'
    assert "data;receita_prevista;receita_recebida;despesas_pagas;saldo" in body
    assert "2026-09-05;800.00;0.00;0.00;0.00" in body
    assert "2026-09-06;0.00;800.00;250.00;550.00" in body


def test_cash_flow_export_xlsx_returns_period_entries(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    student = Student(full_name="Aluno Fluxo Export", status=StudentStatus.ACTIVE)
    receivable = Receivable(
        student=student,
        description="Mensalidade exportacao",
        amount=Decimal("800.00"),
        paid_amount=Decimal("800.00"),
        due_date=date(2026, 9, 5),
        payment_date=date(2026, 9, 6),
        status=PaymentStatus.PAID,
        type=ReceivableType.TUITION,
    )
    payable = Payable(
        description="Fornecedor exportacao",
        category="servicos",
        amount=Decimal("250.00"),
        due_date=date(2026, 9, 6),
        payment_date=date(2026, 9, 6),
        status=PaymentStatus.PAID,
    )
    db.add_all([student, receivable, payable])
    db.commit()

    response = client.get(
        "/api/finance/cash-flow/export.xlsx",
        params={"start_date": "2026-09-05", "end_date": "2026-09-06"},
        headers=auth_headers(client, "finance@example.com"),
    )

    rows = workbook_rows(response)
    flattened = [value for row in rows for value in row if value is not None]
    assert response.status_code == 200
    assert response.headers["content-disposition"] == 'attachment; filename="fluxo_caixa.xlsx"'
    assert "Relatorio de Fluxo de Caixa" in flattened
    assert 800 in flattened
    assert 550 in flattened


def test_dashboard_includes_pending_expenses(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    db.add_all(
        [
            Student(full_name="Aluno Ativo", status=StudentStatus.ACTIVE),
            Payable(
                description="Despesa pendente",
                category="aluguel",
                amount=Decimal("450.00"),
                due_date=date(2026, 9, 20),
                status=PaymentStatus.PENDING,
            ),
        ]
    )
    db.commit()

    response = client.get(
        "/api/dashboard",
        params={"reference_date": "2026-09-15"},
        headers=auth_headers(client, "finance@example.com"),
    )

    assert response.status_code == 200
    assert response.json()["despesas_pendentes_mes"] == "450.00"
    assert response.json()["custo_medio_por_aluno"] == "450.00"


def test_dashboard_comparison_uses_due_payment_and_previous_month_rules(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    student_a = Student(full_name="Aluno Caixa A", status=StudentStatus.ACTIVE)
    student_b = Student(full_name="Aluno Caixa B", status=StudentStatus.ACTIVE)
    db.add_all([student_a, student_b])
    db.flush()
    db.add_all(
        [
            Receivable(
                student_id=student_a.id,
                description="Mensalidade paga setembro",
                amount=Decimal("500.00"),
                paid_amount=Decimal("500.00"),
                due_date=date(2026, 9, 5),
                payment_date=date(2026, 9, 6),
                status=PaymentStatus.PAID,
                type=ReceivableType.TUITION,
            ),
            Receivable(
                student_id=student_a.id,
                description="Mensalidade pendente setembro",
                amount=Decimal("700.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 9, 10),
                status=PaymentStatus.PENDING,
                type=ReceivableType.EXTRA_FEE,
            ),
            Receivable(
                student_id=student_a.id,
                description="Taxa aberta setembro",
                amount=Decimal("300.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 9, 1),
                status=PaymentStatus.PENDING,
                type=ReceivableType.OTHER,
            ),
            Receivable(
                student_id=student_b.id,
                description="Receita recebida em setembro de agosto",
                amount=Decimal("400.00"),
                paid_amount=Decimal("400.00"),
                due_date=date(2026, 8, 20),
                payment_date=date(2026, 9, 2),
                status=PaymentStatus.PAID,
                type=ReceivableType.EXTRA_FEE,
            ),
            Receivable(
                student_id=student_b.id,
                description="Receita cancelada setembro",
                amount=Decimal("200.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 9, 25),
                status=PaymentStatus.CANCELED,
                type=ReceivableType.OTHER,
            ),
            Receivable(
                student_id=student_b.id,
                description="Mensalidade paga agosto",
                amount=Decimal("600.00"),
                paid_amount=Decimal("600.00"),
                due_date=date(2026, 8, 5),
                payment_date=date(2026, 8, 4),
                status=PaymentStatus.PAID,
                type=ReceivableType.EXTRA_FEE,
            ),
            Receivable(
                student_id=student_b.id,
                description="Mensalidade atrasada agosto",
                amount=Decimal("100.00"),
                paid_amount=Decimal("0.00"),
                due_date=date(2026, 8, 10),
                status=PaymentStatus.PENDING,
                type=ReceivableType.OTHER,
            ),
            Payable(
                description="Despesa paga setembro",
                category="operacional",
                amount=Decimal("250.00"),
                due_date=date(2026, 9, 8),
                payment_date=date(2026, 9, 8),
                status=PaymentStatus.PAID,
            ),
            Payable(
                description="Despesa pendente setembro",
                category="operacional",
                amount=Decimal("150.00"),
                due_date=date(2026, 9, 20),
                status=PaymentStatus.PENDING,
            ),
            Payable(
                description="Despesa cancelada setembro",
                category="operacional",
                amount=Decimal("50.00"),
                due_date=date(2026, 9, 21),
                status=PaymentStatus.CANCELED,
            ),
            Payable(
                description="Despesa paga agosto",
                category="operacional",
                amount=Decimal("100.00"),
                due_date=date(2026, 8, 6),
                payment_date=date(2026, 8, 6),
                status=PaymentStatus.PAID,
            ),
            Payable(
                description="Despesa pendente agosto",
                category="operacional",
                amount=Decimal("80.00"),
                due_date=date(2026, 8, 12),
                status=PaymentStatus.PENDING,
            ),
        ]
    )
    db.commit()

    response = client.get(
        "/api/dashboard",
        params={"reference_date": "2026-09-15"},
        headers=auth_headers(client, "finance@example.com"),
    )

    assert response.status_code == 200
    body = response.json()
    assert body["mes_referencia"] == "2026-09"
    assert body["mes_comparacao"] == "2026-08"
    assert body["periodo_inicio"] == "2026-09-01"
    assert body["periodo_fim"] == "2026-09-30"
    assert body["data_referencia"] == "2026-09-30"
    assert body["total_alunos_ativos"] == 2
    assert body["receita_prevista_mes"] == "1500.00"
    assert body["receita_recebida_mes"] == "900.00"
    assert body["despesas_mes"] == "250.00"
    assert body["despesas_pendentes_mes"] == "150.00"
    assert body["saldo_mes"] == "650.00"
    assert body["quantidade_inadimplentes"] == 2
    assert body["valor_total_inadimplente"] == "1100.00"
    assert body["custo_medio_por_aluno"] == "200.00"
    assert body["custo_caixa_por_aluno"] == "125.00"
    assert body["variacao_receita_prevista"] == "400.00"
    assert body["variacao_receita_recebida"] == "300.00"
    assert body["variacao_despesas"] == "150.00"
    assert body["variacao_despesas_pendentes"] == "70.00"
    assert body["variacao_saldo"] == "150.00"
    assert body["variacao_quantidade_inadimplentes"] == 1
    assert body["variacao_valor_inadimplente"] == "1000.00"
    assert body["variacao_custo_medio_por_aluno"] == "110.00"
    assert body["variacao_custo_caixa_por_aluno"] == "75.00"


def test_dashboard_distinguishes_real_cost_from_cash_cost(client, db):
    create_user(db, email="finance@example.com", role=Role.FINANCEIRO)
    db.add_all(
        [
            Student(full_name="Aluno A", status=StudentStatus.ACTIVE),
            Student(full_name="Aluno B", status=StudentStatus.ACTIVE),
            Payable(
                description="Despesa de agosto paga em setembro",
                category="operacional",
                amount=Decimal("300.00"),
                due_date=date(2026, 8, 28),
                payment_date=date(2026, 9, 2),
                status=PaymentStatus.PAID,
            ),
            Payable(
                description="Despesa de setembro pendente",
                category="operacional",
                amount=Decimal("200.00"),
                due_date=date(2026, 9, 20),
                status=PaymentStatus.PENDING,
            ),
        ]
    )
    db.commit()

    response = client.get(
        "/api/dashboard",
        params={"reference_date": "2026-09-15"},
        headers=auth_headers(client, "finance@example.com"),
    )

    assert response.status_code == 200
    body = response.json()
    assert body["despesas_mes"] == "300.00"
    assert body["despesas_pendentes_mes"] == "200.00"
    assert body["custo_caixa_por_aluno"] == "150.00"
    assert body["custo_medio_por_aluno"] == "100.00"

